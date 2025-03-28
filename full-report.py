import json
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

### ========== SCA PART ==========
def merge_sca(old_file, new_file, output_file, key_col="ID"):
    df_old = pd.read_excel(old_file)
    df_new = pd.read_excel(new_file)

    df_old[key_col] = df_old[key_col].astype(str)
    df_new[key_col] = df_new[key_col].astype(str)

    df_old["SourceFile"] = ""
    df_new["SourceFile"] = new_file

    ids_old = set(df_old[key_col])
    df_new_filtered = df_new[~df_new[key_col].isin(ids_old)]

    df_combined = pd.concat([df_old, df_new_filtered], ignore_index=True)

    try:
        wb = load_workbook(output_file)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)
        wb.save(output_file)

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_combined.to_excel(writer, sheet_name="Combined", index=False)

    # Highlight new rows
    wb = load_workbook(output_file)
    ws = wb["Combined"]
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    source_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "SourceFile":
            source_col_idx = idx
            break

    if source_col_idx:
        for row in ws.iter_rows(min_row=2):
            if row[source_col_idx - 1].value:
                for cell in row:
                    cell.fill = green_fill

    wb.save(output_file)
    print("✅ SCA merged & highlighted into sheet: Combined")

### ========== SAST PART ==========
def extract_semgrep_findings(sarif_data, source_label):
    findings = []
    runs = sarif_data.get("runs", [])
    for run in runs:
        for result in run.get("results", []):
            rule_id = result.get("ruleId")
            message = result.get("message", {}).get("text", "")
            level = result.get("level", "")
            locations = result.get("locations", [])
            location_info = locations[0]["physicalLocation"] if locations else {}
            file_path = location_info.get("artifactLocation", {}).get("uri", "")
            region = location_info.get("region", {})
            start_line = region.get("startLine")
            findings.append({
                "RuleID": rule_id,
                "Message": message,
                "Level": level,
                "File": file_path,
                "Line": start_line,
                "SourceFile": source_label
            })
    return findings

def merge_sast(sarif_old_path, sarif_new_path, output_excel_path):
    with open(sarif_old_path, "r", encoding="utf-8") as f:
        sarif_old = json.load(f)
    with open(sarif_new_path, "r", encoding="utf-8") as f:
        sarif_new = json.load(f)

    old_findings = extract_semgrep_findings(sarif_old, "")
    new_findings = extract_semgrep_findings(sarif_new, "semgrep2.sarif")

    def to_key(f):
        return f"{f['RuleID']}::{f['File']}::{f['Line']}"

    old_keys = set(to_key(f) for f in old_findings)
    new_only = [f for f in new_findings if to_key(f) not in old_keys]

    all_findings = old_findings + new_only
    df = pd.DataFrame(all_findings)

    with pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name="SAST", index=False)

    wb = load_workbook(output_excel_path)
    ws = wb["SAST"]

    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    source_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "SourceFile":
            source_col_idx = idx
            break

    if source_col_idx:
        for row in ws.iter_rows(min_row=2):
            if row[source_col_idx - 1].value:
                for cell in row:
                    cell.fill = green_fill

    wb.save(output_excel_path)
    print("✅ SAST merged & highlighted into sheet: SAST")

### ========== MAIN ==========
if __name__ == "__main__":
    merge_sca("unique_vulns_filtered2.xlsx", "unique_vulns_filtered.xlsx", "merged_output.xlsx")
    merge_sast("semgrep.sarif", "semgrep2.sarif", "merged_output.xlsx")
    print("🎉 All reports merged successfully into merged_output.xlsx")