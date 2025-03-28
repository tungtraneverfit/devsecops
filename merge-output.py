import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def merge_and_highlight(old_file, new_file, output_file, key_col="ID"):
    # Load dữ liệu
    df_old = pd.read_excel(old_file)
    df_new = pd.read_excel(new_file)

    # Chuẩn hoá ID để so sánh
    df_old[key_col] = df_old[key_col].astype(str)
    df_new[key_col] = df_new[key_col].astype(str)

    # Đánh dấu file mới bằng Fixed At
    df_old["Fixed At"] = ""
    df_new["Fixed At"] = new_file

    # Loại trùng ID, chỉ giữ bản từ file old nếu trùng
    ids_old = set(df_old[key_col])
    df_new_filtered = df_new[~df_new[key_col].isin(ids_old)]

    # Gộp bảng
    df_combined = pd.concat([df_old, df_new_filtered], ignore_index=True)
    df_combined.to_excel(output_file, index=False, sheet_name="Combined")

    # Tô màu xanh nếu có Fixed At
    wb = load_workbook(output_file)
    ws = wb["Combined"]

    source_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Fixed At":
            source_col_idx = idx
            break

    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    if source_col_idx:
        for row in ws.iter_rows(min_row=2):
            source_val = row[source_col_idx - 1].value
            if source_val:  # Chỉ highlight nếu Fixed At có giá trị
                for cell in row:
                    cell.fill = green_fill

    wb.save(output_file)
    print(f"✅ Done! Output written to {output_file}")

# Example usage
merge_and_highlight("unique_vulns_filtered.xlsx", "unique_vulns_filtered2.xlsx", "merged_output.xlsx")

# Example usage

